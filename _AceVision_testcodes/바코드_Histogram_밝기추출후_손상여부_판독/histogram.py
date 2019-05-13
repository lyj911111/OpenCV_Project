'''
    바코드의 손상 여부를 판단하기위해 검정색과 흰색의 히스토그램 분포도를 판독하여
    바코드 손상여부를 판단과 동시에 어느 부분이 문제가 있는지 마킹을 해줌.
	손상된 부분의 치수값을 엑셀의 그래프로 추출하기 위해 값을 엑셀로저장 (옵션)
'''

import cv2 as cv
import numpy as np

'''
    list 단위로 들어온 값을 col방향으로 엑셀에 저장.
    엑셀로 중앙 히스토그램 분포 그래프를 보기 위함.
'''
import openpyxl     # 엑셀 라이브러리
def writeExcel(list_value):
    # 파일 쓰기 위한 객체 생성.
    write_wb = openpyxl.Workbook()
    write_ws = write_wb.active

    # 엑셀에 값 쓰기
    for i in range(0, len(list_value)):
        string = list_value[i]
        write_ws.cell(i+1, 1, i+1)          # 1번째 column - 인덱스값 x값
        write_ws.cell(i+1, 2, string)       # 2번째 column - 치수값 y값

    # 편집된 엑셀파일을 저장.
    write_wb.save('./test.xlsx')


'''
원본 이미지에 마스크를 씌우고 그 부분에서의 밝기에 대한 histogram을 찾는 함수. (밝기에 대한 분포도)
    param 1 : 히스토그램을 구할 원본이미지
    param 2~4: (x1,y1) 부터 (x2, y2) 의 이미지 마스크를 씌워서 구분함
'''
def mask_histogram(img, x1,y1,x2,y2):
    h = np.zeros((img.shape[0], 256), dtype=np.uint8)   # 히스토그램을 만들 공간 생성.

    # mask생성
    mask = np.zeros(img.shape[:2], np.uint8)
    mask[y1:y2, x1:x2] = 255                 # 검정 바탕(픽셀값0)에 [세로 x 가로]의 부분만 픽셀을 255으로 마스크 씌움

    # 이미지에 mask가 적용된 결과
    masked_img = cv.bitwise_and(img, img, mask=mask)

    # Histogram: [이미지], [ 채널: [0]그래이스케일/ 컬러 경우, BGR채널[0][1][2] ], 전체이미지 : None , x축범위, 막대의 갯수
    hist_item = cv.calcHist([img], [0], mask, [256], [0, 256])

    cv.normalize(hist_item, hist_item, 0, 255, cv.NORM_MINMAX)
    hist = np.int32(np.around(hist_item))
    for x, y in enumerate(hist):
        cv.line(h, (x, 0 + 10), (x, y + 10), (255, 255, 255))

    cv.line(h, (0, 0 + 10), (0, 5), (255, 255, 255))
    cv.line(h, (255, 0 + 10), (255, 5), (255, 255, 255))
    y = np.flipud(h)

    result = np.hstack((masked_img, y))  # 이미지와 히스토그램을 붙이기

    # histogram의 y값, 마스크씌운 이미지를 리턴.
    return hist_item, result


def draw_histogram(img):
    h = np.zeros((img.shape[0], 256), dtype=np.uint8)

    # Histogram: [이미지], [ 채널: [0]그래이스케일/ 컬러 경우, BGR채널[0][1][2] ], 전체이미지 : None , x축범위, 막대의 갯수
    hist_item = cv.calcHist([img], [0], None, [200], [0, 200])


    cv.normalize(hist_item,hist_item,0,255,cv.NORM_MINMAX)
    hist = np.int32(np.around(hist_item))
    for x, y in enumerate(hist):
        cv.line(h, (x, 0+10), (x, y+10), (255, 255, 255))

    cv.line(h, (0, 0 + 10), (0, 5), (255, 255, 255) )
    cv.line(h, (255, 0 + 10), (255, 5), (255, 255, 255))
    y = np.flipud(h)

    result = np.hstack((img, y))  # 이미지와 히스토그램을 붙이기

    # histogram의 y값, 마스크씌운 이미지를 리턴.
    return hist_item, result

# 이미지 불러오기
img = cv.imread('./bacodeTest.PNG', cv.IMREAD_COLOR)
img = cv.resize(img,(1260,960))
gray = cv.cvtColor(img, cv.COLOR_BGR2GRAY)

##################################################### 사용자 지정 영역.
# 스캔시작좌표(x, y) / 스캔할 범위 지정. (x축 가로넓이)
x_coord = 100
y_coord = 295
scan_range = 640

# 스캔 폭 지정. <= 스캔폭을 잘 지정해야 명확히 인식함. (8~15)추천값
scan_width = 15

# 스캔의 정밀도 (스캔 량)
scan_density = 1

# 스캔할 높이지정.
scan_height = 35
######################################################

# 어디를 스캔할지 미리 보여줌. 키보드를 누르면 시작.
y, result = mask_histogram(gray, x_coord, y_coord, scan_range, y_coord + scan_height)
cv.imshow('result', result)
cv.waitKey(0)

# 왼쪽에서부터 오른쪽으로 차례대로 스캔하면서 감.
sum_mid_list = []
histo_list = []
mark_list = []
sum_mid = 0
for j in range(0, scan_range - (x_coord + scan_width), scan_density):
    # 마스크 씌운 이미지의 히스토그램
    y, result = mask_histogram(gray, x_coord + j, y_coord, (x_coord + scan_width) + j, y_coord + scan_height)
    result_copy = cv.cvtColor(result, cv.COLOR_GRAY2RGB)

    #print(y)    # 그래프의 막대 좌표에 대한 정보

    # 0 ~ 255 픽셀 히스토그램에서 77 ~ 116 구간 (회색이 가장 높은 구간)의 값을 모두 합함.
    histo_list = []

    for i in range(len(y)):
        print("y인덱스 %d : " %i, y[i][0])
        if i > 77 and i < 116:
            histo_list.append(y[i][0])
            sum_mid = sum(histo_list)       # 회색성분이 높은 구간의 합
            sum_mid_list.append(sum_mid)

    print("중앙합리스트", sum_mid_list)

    # 회색 성분의 합의 3000을 넘어갈때 바코드 손상으로 판단, (핵심 부분.)
    if sum_mid > 3000:
        mark_list.append(x_coord + j)
    print("중앙회색성분 합:", sum_mid)

    # 손상된 부분 마킹.
    for i in range(len(mark_list)):
        cv.circle(result_copy, (int(mark_list[i] + (scan_width/2)), y_coord-30), 10, (255, 0, 255), -1)

    # 스캔하고 있는 부분 마커.
    cv.circle(result_copy, (int(x_coord + j + (scan_width/2)), y_coord-30), 10, (255, 0, 255), -1)
    cv.imshow('result', result_copy)
    cv.waitKey(1)

# 다된 결과 표시, 전체 마스크된 부분, 컬러 마크된 부분
y, result = mask_histogram(gray, x_coord, y_coord, scan_range, y_coord + scan_height)
result_copy = cv.cvtColor(result, cv.COLOR_GRAY2RGB)
for i in range(len(mark_list)): # 마크된 원 표시
    cv.circle(result_copy, (int(mark_list[i] + (scan_width / 2)), y_coord - 30), 5, (255, 0, 255), -1)
cv.imshow('result', result_copy)

cv.waitKey(0)

# 중앙 분포도 값을 그래프로 보기위해 수치를 엑셀로 저장 (엑셀 그래프 확인용)
writeExcel(sum_mid_list)