import openpyxl     # pip install openpyxl

'''
    엑셀 파일 쓰기.
'''
# 파일 쓰기 위한 객체 생성.
write_wb = openpyxl.Workbook()
write_ws = write_wb.active

# (col, row 인자값) 에 입력
write_ws = write_wb.active
write_ws['A1'] = '수량'
write_ws['B1'] = '가격'
write_ws['C1'] = '금액'

# 행(Row) 단위로 추가
write_ws.append([25, 5000, 125000])
write_ws.append([30, 5500, 60000])  # 열(row) 단위로 아래로 내려감.
write_ws.append([35, 8000, 9000])

for i in range(5):
    write_ws.append([i+1, i+2, i+3])


# 셀 단위로 추가 col2, row5에 '셀추가'를 입력.
write_ws.cell(5, 5, '셀추가')

# 편집된 엑셀파일을 저장.
write_wb.save('./test.xlsx')


'''
    엑셀 파일 읽기.
'''
# 파일 쓰기 위한 객체 생성.
load_wb = openpyxl.load_workbook("./test.xlsx", data_only=True)
# sheet 이름 불러오기
load_ws = load_wb['Sheet']

# 셀주소값에 접근 후 출력
print(load_ws['A1'].value)

# 셀좌표값에 접근 후 출력
print(load_ws.cell(2,2).value)

# 셀 범위 지정하여 차례대로 출력
get_cells = load_ws['A1':'C3']
for row in get_cells:
    for cell in row:
        print(cell.value)


# 모든 행과 열을 출력
print("\n=== 모든 행과 열 출력 ===")
all_values = []
for row in load_ws.rows:
    row_value = []
    for cell in row:
        row_value.append(cell.value)
    all_values.append(row_value)

print(all_values)
